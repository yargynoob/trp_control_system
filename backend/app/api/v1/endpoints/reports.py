"""Report endpoints."""

from typing import List
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
import os
import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.db import get_db
from app.models.report import Report
from app.models.user import User
from app.models.defect import Defect, DefectStatus, Priority
from app.models.project import Project
from app.core.deps import get_current_user, user_has_role_in_project
from app.schemas.report import ReportCreate, ReportResponse, ReportList, ReportFormat as SchemaReportFormat

router = APIRouter()

REPORTS_DIR = "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)


@router.post("/", response_model=ReportResponse, status_code=status.HTTP_201_CREATED)
def create_report(
    report_data: ReportCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new report. Only supervisors can create reports."""
    # Determine if single or multi-project report
    is_multi_project = report_data.project_ids is not None
    project_list = report_data.project_ids if is_multi_project else [report_data.project_id]
    
    # Verify all projects exist and user is supervisor in all of them
    projects = []
    for proj_id in project_list:
        project = db.query(Project).filter(Project.id == proj_id).first()
        if not project:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Project {proj_id} not found"
            )
        
        is_supervisor = user_has_role_in_project(current_user.id, proj_id, ['supervisor'], db)
        if not is_supervisor:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Only supervisors can create reports. You are not a supervisor in project {proj_id}"
            )
        projects.append(project)
    
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_extension = 'xlsx' if report_data.format == SchemaReportFormat.EXCEL else 'csv'
        
        if is_multi_project:
            project_ids_str = "_".join(map(str, sorted(project_list)))
            filename = f"report_multi_{project_ids_str}_{timestamp}.{file_extension}"
        else:
            filename = f"report_{report_data.project_id}_{timestamp}.{file_extension}"
        
        file_path = os.path.join(REPORTS_DIR, filename)
        
        # Get defects from all projects
        defects_query = db.query(
            Defect.id,
            Defect.project_id,
            Defect.title,
            Defect.description,
            Defect.location,
            DefectStatus.display_name.label('status'),
            Priority.display_name.label('priority'),
            User.first_name,
            User.last_name,
            Defect.due_date,
            Defect.created_at,
            Defect.updated_at,
            Project.name.label('project_name')
        ).join(
            DefectStatus, Defect.status_id == DefectStatus.id
        ).join(
            Priority, Defect.priority_id == Priority.id
        ).join(
            Project, Defect.project_id == Project.id
        ).outerjoin(
            User, Defect.assignee_id == User.id
        ).filter(
            Defect.project_id.in_(project_list)
        )
        
        defects_data = defects_query.all()
        
        # Generate report
        if is_multi_project:
            project_names = ", ".join([p.name for p in projects])
            report_title = f"Multi-Project Report: {project_names}"
        else:
            report_title = projects[0].name
        
        if report_data.format == SchemaReportFormat.CSV:
            _generate_csv_report(file_path, defects_data, is_multi_project)
        else:  # EXCEL
            _generate_excel_report(file_path, defects_data, report_title, is_multi_project)
        
        file_size = os.path.getsize(file_path)
        
        db_report = Report(
            project_id=report_data.project_id if not is_multi_project else None,
            project_ids=report_data.project_ids if is_multi_project else None,
            created_by=current_user.id,
            title=report_data.title,
            description=report_data.description,
            format=report_data.format.value,  
            file_path=file_path,
            file_size=file_size
        )
        db.add(db_report)
        db.commit()
        db.refresh(db_report)
        
        creator_name = f"{current_user.first_name} {current_user.last_name}" if current_user.first_name else current_user.username
        
        return {
            "id": db_report.id,
            "project_id": db_report.project_id,
            "project_ids": db_report.project_ids,
            "created_by": db_report.created_by,
            "creator_name": creator_name,
            "title": db_report.title,
            "description": db_report.description,
            "format": db_report.format,
            "file_path": db_report.file_path,
            "file_size": db_report.file_size,
            "created_at": db_report.created_at
        }
        
    except Exception as e:
        db.rollback()
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create report: {str(e)}"
        )


@router.get("/project/{project_id}", response_model=ReportList)
def get_project_reports(
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all reports accessible to the user for a specific project or all projects."""
    # Special handling: if project_id is 0, get reports from ALL projects where user is supervisor
    if project_id == 0:
        # Get all projects where user is supervisor
        from app.models.role import UserRole, Role
        user_supervisor_projects = db.query(UserRole.project_id).join(
            Role, UserRole.role_id == Role.id
        ).filter(
            UserRole.user_id == current_user.id,
            Role.name == "supervisor"
        ).all()
        
        supervisor_project_ids = [up.project_id for up in user_supervisor_projects]
        
        if not supervisor_project_ids:
            return {"reports": [], "total": 0}
        
        # Get all reports
        all_reports = db.query(Report, User).join(
            User, Report.created_by == User.id
        ).order_by(Report.created_at.desc()).all()
        
        # Filter reports based on access rights
        result = []
        for report, creator in all_reports:
            # Determine which projects this report covers
            report_project_ids = report.project_ids if report.project_ids else [report.project_id]
            
            # User can see this report only if they are supervisor in ALL projects
            if all(pid in supervisor_project_ids for pid in report_project_ids):
                creator_name = f"{creator.first_name} {creator.last_name}" if creator.first_name else creator.username
                result.append({
                    "id": report.id,
                    "project_id": report.project_id,
                    "project_ids": report.project_ids,
                    "created_by": report.created_by,
                    "creator_name": creator_name,
                    "title": report.title,
                    "description": report.description,
                    "format": report.format,
                    "file_path": report.file_path,
                    "file_size": report.file_size,
                    "created_at": report.created_at
                })
        
        return {
            "reports": result,
            "total": len(result)
        }
    
    # Single project view
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    is_supervisor = user_has_role_in_project(current_user.id, project_id, ['supervisor'], db)
    is_manager = user_has_role_in_project(current_user.id, project_id, ['manager'], db)
    
    if not (is_supervisor or is_manager):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only supervisors and managers can view reports"
        )
    
    # Get reports for this project (including multi-project reports that include this project)
    from sqlalchemy import any_
    reports = db.query(Report, User).join(
        User, Report.created_by == User.id
    ).filter(
        (Report.project_id == project_id) | 
        (project_id == any_(Report.project_ids))
    ).order_by(Report.created_at.desc()).all()
    
    result = []
    for report, creator in reports:
        # For multi-project reports, check if user has access to ALL projects in the report
        if report.project_ids:
            # This is a multi-project report
            has_access_to_all = True
            for proj_id in report.project_ids:
                proj_is_supervisor = user_has_role_in_project(current_user.id, proj_id, ['supervisor'], db)
                proj_is_manager = user_has_role_in_project(current_user.id, proj_id, ['manager'], db)
                if not (proj_is_supervisor or proj_is_manager):
                    has_access_to_all = False
                    break
            
            # Skip this report if user doesn't have access to all projects
            if not has_access_to_all:
                continue
        
        creator_name = f"{creator.first_name} {creator.last_name}" if creator.first_name else creator.username
        result.append({
            "id": report.id,
            "project_id": report.project_id,
            "project_ids": report.project_ids,
            "created_by": report.created_by,
            "creator_name": creator_name,
            "title": report.title,
            "description": report.description,
            "format": report.format,
            "file_path": report.file_path,
            "file_size": report.file_size,
            "created_at": report.created_at
        })
    
    return {
        "reports": result,
        "total": len(result)
    }


@router.get("/{report_id}/download")
def download_report(
    report_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found"
        )
    
    # Check access rights based on report type
    has_access = False
    
    if report.project_ids:
        # Multi-project report - check if user has access to at least one project
        for proj_id in report.project_ids:
            is_supervisor = user_has_role_in_project(current_user.id, proj_id, ['supervisor'], db)
            is_manager = user_has_role_in_project(current_user.id, proj_id, ['manager'], db)
            if is_supervisor or is_manager:
                has_access = True
                break
    elif report.project_id:
        # Single project report
        is_supervisor = user_has_role_in_project(current_user.id, report.project_id, ['supervisor'], db)
        is_manager = user_has_role_in_project(current_user.id, report.project_id, ['manager'], db)
        has_access = is_supervisor or is_manager
    
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only supervisors and managers can download reports"
        )
    
    if not os.path.exists(report.file_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report file not found"
        )
    
    filename = os.path.basename(report.file_path)
    media_type = "text/csv" if report.format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return FileResponse(
        path=report.file_path,
        filename=filename,
        media_type=media_type
    )


@router.get("/{report_id}/chart-data")
def get_report_chart_data(
    report_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get chart data for a report (defects created in the last 30 days)."""
    from datetime import timedelta
    
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found"
        )
    
    project_ids = report.project_ids if report.project_ids else [report.project_id]
    
    for proj_id in project_ids:
        is_supervisor = user_has_role_in_project(current_user.id, proj_id, ['supervisor'], db)
        is_manager = user_has_role_in_project(current_user.id, proj_id, ['manager'], db)
        
        if not (is_supervisor or is_manager):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied to report projects"
            )
    
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    all_dates = []
    current = start_date.date()
    while current <= end_date.date():
        all_dates.append(current.strftime('%Y-%m-%d'))
        current += timedelta(days=1)
    
    result = {}
    
    for proj_id in project_ids:
        project = db.query(Project).filter(Project.id == proj_id).first()
        project_name = project.name if project else f"Project {proj_id}"
        
        defects_by_day = db.query(
            func.date(Defect.created_at).label('date'),
            func.count(Defect.id).label('count')
        ).filter(
            Defect.project_id == proj_id,
            Defect.created_at >= start_date,
            Defect.created_at <= end_date
        ).group_by(
            func.date(Defect.created_at)
        ).all()
        
        defects_dict = {item.date.strftime('%Y-%m-%d'): item.count for item in defects_by_day}
        
        data = [defects_dict.get(date, 0) for date in all_dates]
        
        result[proj_id] = {
            "project_id": proj_id,
            "project_name": project_name,
            "data": data
        }
    
    return {
        "dates": all_dates,
        "projects": list(result.values())
    }


def _generate_csv_report(file_path: str, defects_data, is_multi_project=False):
    """Generate CSV report with optional project column for multi-project reports."""
    with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        
        headers = ['ID']
        if is_multi_project:
            headers.append('Проект')
        headers.extend([
            'Название',
            'Описание',
            'Местоположение',
            'Статус',
            'Приоритет',
            'Исполнитель',
            'Срок выполнения',
            'Дата создания',
            'Дата обновления'
        ])
        writer.writerow(headers)
        
        for defect in defects_data:
            assignee = f"{defect.first_name} {defect.last_name}" if defect.first_name else "Не назначен"
            
            row = [defect.id]
            if is_multi_project:
                row.append(defect.project_name)
            row.extend([
                defect.title,
                defect.description,
                defect.location,
                defect.status,
                defect.priority,
                assignee,
                defect.due_date.strftime('%d.%m.%Y') if defect.due_date else '',
                defect.created_at.strftime('%d.%m.%Y %H:%M'),
                defect.updated_at.strftime('%d.%m.%Y %H:%M') if defect.updated_at else ''
            ])
            writer.writerow(row)


def _generate_excel_report(file_path: str, defects_data, report_title: str, is_multi_project=False):
    """Generate Excel report with optional project column for multi-project reports."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Defects"
    
    num_cols = 11 if is_multi_project else 10
    last_col_letter = chr(64 + num_cols)
    ws.merge_cells(f'A1:{last_col_letter}1')
    title_cell = ws['A1']
    title_cell.value = f"Отчет по дефектам - {report_title}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 5
    
    headers = ['ID']
    if is_multi_project:
        headers.append('Проект')
    headers.extend([
        'Название', 'Описание', 'Местоположение', 'Статус',
        'Приоритет', 'Исполнитель', 'Срок выполнения', 'Дата создания', 'Дата обновления'
    ])
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, defect in enumerate(defects_data, 4):
        assignee = f"{defect.first_name} {defect.last_name}" if defect.first_name else "Не назначен"
        
        col = 1
        ws.cell(row=row_num, column=col, value=defect.id)
        col += 1
        
        if is_multi_project:
            ws.cell(row=row_num, column=col, value=defect.project_name)
            col += 1
        
        ws.cell(row=row_num, column=col, value=defect.title)
        ws.cell(row=row_num, column=col+1, value=defect.description)
        ws.cell(row=row_num, column=col+2, value=defect.location)
        ws.cell(row=row_num, column=col+3, value=defect.status)
        ws.cell(row=row_num, column=col+4, value=defect.priority)
        ws.cell(row=row_num, column=col+5, value=assignee)
        ws.cell(row=row_num, column=col+6, value=defect.due_date.strftime('%d.%m.%Y') if defect.due_date else '')
        ws.cell(row=row_num, column=col+7, value=defect.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row_num, column=col+8, value=defect.updated_at.strftime('%d.%m.%Y %H:%M') if defect.updated_at else '')
    
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = ws.cell(row=3, column=col_num).column_letter
        
        header_length = len(str(headers[col_num - 1]))
        if header_length > max_length:
            max_length = header_length
        
        for row_num in range(4, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
